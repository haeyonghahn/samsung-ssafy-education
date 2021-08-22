import urllib.request

from bs4 import BeautifulSoup
from openpyxl import Workbook   # 엑셀 라이브러리(pip install openpyxl)
# from elice_utils import EliceUtils

# elice_utils = EliceUtils()


def main():
    
  wb = Workbook() # 엑셀 라이브러리를 쓸 것이다라고 정의
  
  sheet1 = wb.active  # wb에서 기본 sheet를 활성화
  file_name = 'sample.xlsx'   # 시트 제목 지정

  url = "http://www.newsis.com/eco/list/?cid=10400&scid=10404"
  req = urllib.request.Request(url)
  sourcecode = urllib.request.urlopen(url).read()
  soup = BeautifulSoup(sourcecode, "html.parser")

  articles = []

  for i in soup.find_all("strong", class_="title"):
      articles.append(i.get_text())
  #print(articles)
  
  
  # range의 범위를 0부터 시작하지 않은 이유는 엑셀 시작이 1부터 시작함
  for row_index in range(1, len(articles) + 1):
      sheet1.cell(row=row_index, column=1).value = row_index
      # row_index-1로 시작하는 이유는 articles 리스트 인덱스가 0부터 시작하기 때문에
      sheet1.cell(row=row_index, column=2).value = articles[row_index-1]

  wb.save(filename=file_name) # 엑셀 파일로 세이브
  # elice_utils.send_file(file_name)

if __name__ == "__main__":
  main()